import pandas as pd
import numpy as np
import openpyxl

df = pd.DataFrame()
path = #path to your workbook

def write_to_excel(df: pd.DataFrame, path_to_workbook: str, sheet_name: str = 'data') -> None:
    """Inserts data into an existing workbook, without affecting other sheets.

    Args:
        df (pd.DataFrame): dataframe to format into a sheet
        path_to_workbook (str): path to workbook
        sheet_name (str, optional): sheet to insert data into. Defaults to 'data'.
    """
    book = openpyxl.load_workbook(path)
    
    if len(book.sheetnames) == 1:
        book.create_sheet('qwerty')
    
    try:
        del book[str(sheet_name)]
        book.save(path)
    except KeyError:
        pass
    
    if 'qwerty' in book.sheetnames:
        del book['qwerty']
    
    writer = pd.ExcelWriter(path, engine = 'openpyxl')
    writer.book = book
    
    df.to_excel(writer, sheet_name = 'data')
    writer.save()
    writer.close()
    
write_to_excel(df, path)

